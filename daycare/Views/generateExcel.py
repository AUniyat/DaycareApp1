__author__ = 'ALEX'
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook
import os
from daycare.Database.DatabaseSqlite3 import *

def openExcel():
    wb = Workbook()

    dest_filename = 'family_excel.xlsx'

    ws1 = wb.active
    ws1.title = "Families"

    #Get all the saved families
    db = Database()
    families = db.getAllFamilies()
    allFamilies = []
    for tuple in families:
            family = Family(tuple[0], tuple[1], tuple[2], tuple[3], tuple[4])
            allFamilies.append(family)

    #Pulling the last names from the retrieved family list
    lastNameList = []
    for stuff in allFamilies:
        stuff2 = str(stuff).split()[2]
        lastNameList.append(stuff2)

    #a list made to hold the months of the year
    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
              'August', 'September', 'October', 'November', 'December']

    #Setting the top row equal to the months of the year
    for row in ws1.iter_rows('B1:M1'):
        counter=0
        for cell in row:
          cell.value = months[counter]
          counter=counter+1


    #Setting column A equal to the first name of each family
    for col in range(1, 2):
        test=0
        for row in range(2, 2 + len(lastNameList)):
            test2 = lastNameList[test]
            ws1.cell(column=col, row=row, value=test2)
            test = test + 1


    #Saving the excel file
    wb.save(filename = dest_filename)
    #Open the saved excel file
    os.startfile("family_excel.xlsx")

